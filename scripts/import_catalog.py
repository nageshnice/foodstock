import argparse
import asyncio
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from app.database.session import AsyncSessionFactory, close_database
from app.models.domain import Brand, Category, Product, ProductVariant, Region, Vendor
from app.utils.text import slugify

REGIONS = [
    "Japanese",
    "Korean",
    "Thai",
    "Chinese",
    "Exotics",
    "Spices",
    "Middle Eastern",
    "Singapore & Malaysian",
]
SHEET_REGION = {
    "AL AMEERA": "Middle Eastern",
    "Thai Products": "Thai",
    "Singapore & Malaysian Products": "Singapore & Malaysian",
    "Japanese Products": "Japanese",
}
SHEET_BRAND = {
    "First Choice": "First Choice",
    "AL AMEERA": "Al Ameera",
}

REGION_KEYWORDS = {
    "Japanese": (
        "udon",
        "soba",
        "ramen",
        "sushi",
        "nori",
        "wasabi",
        "hondashi",
        "miso",
        "togarashi",
        "wakame",
        "tempura",
        "karashi",
        "golden curry",
    ),
    "Korean": ("gochujang", "kimchee", "kimchi", "korean"),
    "Thai": ("jasmine rice", "lychee", "water chestnuts", "straw mushroom"),
    "Chinese": (
        "bamboo",
        "bean curd",
        "black bean",
        "chinkiang",
        "pixian",
        "chinese",
        "preserved vegetable",
        "black fungus",
    ),
    "Spices": (
        "sesame",
        "chilly",
        "chilli",
        "pepper",
        "starch",
        "charcoal",
        "agar",
        "mustard",
    ),
    "Exotics": (
        "olive",
        "tomato",
        "peanut",
        "vanilla",
        "butterfly",
        "pine nuts",
        "black rice",
        "red rice",
        "cous cous",
        "jalapenos",
        "kalamata",
        "maraschino",
        "mock duck",
        "barbecue",
    ),
}


async def get_or_create(
    session: object, model: type[object], name: str, **values: object
) -> object:
    existing = await session.scalar(select(model).where(model.name == name))
    if existing:
        return existing
    item = model(name=name, **values)
    session.add(item)
    await session.flush()
    return item


def product_rows(path: Path) -> list[tuple[str, int, str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[tuple[str, int, str, str]] = []
    for sheet in workbook.worksheets:
        for number, name, size in sheet.iter_rows(min_row=3, values_only=True):
            if number and name:
                rows.append(
                    (sheet.title, int(number), str(name).strip(), str(size or "Not specified"))
                )
    workbook.close()
    return rows


def split_sizes(value: str) -> list[str]:
    if " / " not in value:
        return [value]
    return [part.strip() for part in value.split(" / ") if part.strip()]


def infer_region(sheet: str, product_name: str) -> str:
    if sheet in SHEET_REGION:
        return SHEET_REGION[sheet]
    normalized = product_name.lower()
    for region, keywords in REGION_KEYWORDS.items():
        if any(keyword in normalized for keyword in keywords):
            return region
    return "Exotics"


async def import_catalog(path: Path) -> None:
    rows = product_rows(path)
    async with AsyncSessionFactory() as session:
        regions = {
            name: await get_or_create(
                session,
                Region,
                name,
                slug=slugify(name),
                description=f"Imported {name} pantry products",
            )
            for name in REGIONS
        }
        category = await get_or_create(
            session,
            Category,
            "Imported Pantry",
            slug="imported-pantry",
            description="Catalog imported from the Faridi Impex workbook",
        )
        vendor = await get_or_create(session, Vendor, "Faridi Impex Pvt. Ltd.", is_active=True)
        brands = {
            name: await get_or_create(session, Brand, name, slug=slugify(name))
            for name in {"First Choice", "Al Ameera", "Faridi Impex"}
        }

        imported = 0
        updated = 0
        for sheet, number, name, size in rows:
            sheet_code = slugify(sheet).replace("-products", "")[:12].upper()
            sku = f"FI-{sheet_code}-{number:03d}"
            brand_name = SHEET_BRAND.get(sheet, "Faridi Impex")
            region_name = infer_region(sheet, name)
            product = await session.scalar(
                select(Product)
                .where(Product.sku == sku)
                .where(Product.source_section == sheet)
                .options(selectinload(Product.variants))
            )
            if product:
                product.name = name
                product.source_section = sheet
                product.region_id = regions[region_name].id
                product.category_id = category.id
                product.brand_id = brands[brand_name].id
                product.vendor_id = vendor.id
                updated += 1
            else:
                product = Product(
                    sku=sku,
                    name=name,
                    slug=f"{slugify(name)}-{slugify(sheet)}-{number}",
                    description=None,
                    source_section=sheet,
                    tax_rate=5,
                    is_active=False,
                    region_id=regions[region_name].id,
                    category_id=category.id,
                    brand_id=brands[brand_name].id,
                    vendor_id=vendor.id,
                )
                session.add(product)
                await session.flush()
                imported += 1

            existing_sizes = {variant.size.lower() for variant in product.variants}
            product.variants = [
                *product.variants,
                *[
                    ProductVariant(
                        size=variant_size,
                        price=0,
                        stock_quantity=0,
                        is_active=False,
                    )
                    for variant_size in split_sizes(size)
                    if variant_size.lower() not in existing_sizes
                ],
            ]
        await session.commit()
    await close_database()
    print(
        f"Imported {imported} draft products and updated {updated} existing products "
        f"from {len(rows)} spreadsheet rows."
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="Import the Faridi Impex Excel catalog")
    parser.add_argument("xlsx_path", type=Path)
    args = parser.parse_args()
    if not args.xlsx_path.is_file():
        raise FileNotFoundError(args.xlsx_path)
    asyncio.run(import_catalog(args.xlsx_path))


if __name__ == "__main__":
    main()
